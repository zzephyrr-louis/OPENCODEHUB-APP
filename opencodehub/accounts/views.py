from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import CreateView
from .forms import CustomUserCreationForm, CustomLoginForm
from .models import Project, ProjectFile, Comment, ProjectVersion
from django.shortcuts import get_object_or_404
from django.db.models import Q
from accounts.models import User
from django.db import transaction
import time
import os

# Helper function to create version
def create_project_version(project, user, action='manual', description=''):
    """Create a new version snapshot for a project"""
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            with transaction.atomic():
                # Lock and refresh the project
                locked_project = Project.objects.select_for_update().get(id=project.id)
                
                # Mark all existing versions as not latest
                ProjectVersion.objects.filter(
                    project=locked_project,
                    is_latest=True
                ).update(is_latest=False)
                
                # Get latest version with explicit ordering and locking
                latest_version = ProjectVersion.objects.select_for_update().filter(
                    project=locked_project
                ).order_by('-id').first()
                
                if latest_version:
                    try:
                        current_num = int(latest_version.version_number.replace('v', ''))
                        version_number = f"v{current_num + 1}"
                    except:
                        version_number = "v1"
                else:
                    version_number = "v1"
                
                # Create the version - it will be latest
                version = ProjectVersion.objects.create(
                    project=locked_project,
                    version_number=version_number,
                    description=description,
                    action=action,
                    created_by=user,
                    is_latest=True  # Mark as latest
                )
                
                # Create snapshot of current files
                version.create_files_snapshot()
                
                return version
                
        except Exception as e:
            if 'duplicate' in str(e).lower() or 'unique constraint' in str(e).lower():
                if attempt < max_retries - 1:
                    time.sleep(0.1)
                    continue
                else:
                    import datetime
                    timestamp = datetime.datetime.now().strftime('%H%M%S%f')
                    version_number = f"v{timestamp}"
                    
                    with transaction.atomic():
                        ProjectVersion.objects.filter(
                            project=project,
                            is_latest=True
                        ).update(is_latest=False)
                        
                        version = ProjectVersion.objects.create(
                            project=project,
                            version_number=version_number,
                            description=description,
                            action=action,
                            created_by=user,
                            is_latest=True
                        )
                        version.create_files_snapshot()
                        return version
            else:
                raise
    
    return None

def landing(request):
    """Landing page view"""
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'accounts/landing.html')

class CustomRegisterView(CreateView):
    """Registration view"""
    form_class = CustomUserCreationForm
    template_name = 'accounts/register.html'
    success_url = reverse_lazy('login')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Account created successfully! You can now log in.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

class CustomLoginView(LoginView):
    """Login view"""
    form_class = CustomLoginForm
    template_name = 'accounts/login.html'
    redirect_authenticated_user = True
    
    def get_success_url(self):
        return reverse_lazy('home')

    def form_valid(self, form):
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Invalid credentials. Please try again.')
        return super().form_invalid(form)

@login_required
def home(request):
    """Home/Dashboard page view"""
    # Get user's projects
    projects = Project.objects.filter(
        owner=request.user,
        is_deleted=False  
    ).order_by('-updated_at')

    # Get recent activities (last 5 projects updated)
    recent_activities = []
    for project in projects[:5]:
        recent_activities.append({
            'time': f"{project.updated_at.strftime('%H:%M %p')}",
            'project_name': project.title,
            'description': f'Updated {project.updated_at.strftime("%B %d, %Y")}'
        })
    
    context = {
        'projects': projects,
        'recent_activities': recent_activities,
    }
    
    return render(request, 'accounts/home.html', context)

def custom_logout(request):
    """Custom logout view"""
    logout(request)
    return redirect('landing')
    

@login_required
def create_project(request):
    """Create new project"""
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        is_public = request.POST.get('is_public') == 'on'
        
        project = Project.objects.create(
            title=title,
            description=description,
            owner=request.user,
            is_public=is_public
        )

        # CREATE INITIAL VERSION
        create_project_version(
            project=project,
            user=request.user,
            action='created',
            description=f'Project "{title}" created'
        )

        messages.success(request, f'Project "{title}" created successfully!')
        return redirect('project_detail', project_id=project.id)
    
    return render(request, 'accounts/create_project.html')

@login_required
def project_detail(request, project_id):
    """View project details"""
    project = get_object_or_404(Project, id=project_id)
    
    # Check permissions - allow owner, shared users, or public projects
    if not project.is_public and project.owner != request.user and request.user not in project.shared_with.all():
        messages.error(request, 'You do not have permission to view this project.')
        return redirect('home')
    
    # Handle comment posting
    if request.method == 'POST':
        comment_text = request.POST.get('text')
        if comment_text:
            Comment.objects.create(
                project=project,
                author=request.user,
                content=comment_text
            )

            # CREATE VERSION FOR COMMENT
            create_project_version(
                project=project,
                user=request.user,
                action='comment_added',
                description=f'Comment added by {request.user.username}'
            )

            messages.success(request, 'Comment added successfully!')
            return redirect('project_detail', project_id=project.id)
    
    files = project.files.all()
    comments = project.comments.all()
    versions = project.versions.all()[:10]  # Latest 5 versions
    
    return render(request, 'accounts/project_detail.html', {
        'project': project,
        'files': files,
        'comments': comments,
        'versions': versions
    })

@login_required
def upload_file(request, project_id):
    """Upload file to project with validation"""
    from django.conf import settings
    
    project = get_object_or_404(Project, id=project_id)
    
    # Check if user is owner or has access
    if project.owner != request.user and request.user not in project.shared_with.all():
        messages.error(request, 'You do not have permission to upload files to this project.')
        return redirect('project_detail', project_id=project.id)
    
    if request.method == 'POST':
        files = request.FILES.getlist('files')
        
        # Validation constants from settings (with fallback defaults)
        MAX_FILE_SIZE = getattr(settings, 'MAX_FILE_SIZE', 50 * 1024 * 1024)  # 50MB default
        BLOCKED_EXTENSIONS = getattr(settings, 'BLOCKED_FILE_EXTENSIONS', 
            ['.exe', '.bat', '.sh', '.cmd', '.com', '.app', '.dmg', '.deb', '.rpm'])
        
        successful_uploads = 0
        failed_uploads = []
        uploaded_files_names = [] 
        
        for file in files:
            file_name = file.name
            file_size = file.size
            file_extension = os.path.splitext(file_name)[1].lower()
            
            # Security check - block dangerous file types
            if file_extension in BLOCKED_EXTENSIONS:
                failed_uploads.append({
                    'name': file_name,
                    'reason': f'File type {file_extension} is not allowed for security reasons.'
                })
                continue
            
            # File size check
            if file_size > MAX_FILE_SIZE:
                size_mb = file_size / (1024 * 1024)
                failed_uploads.append({
                    'name': file_name,
                    'reason': f'File size ({size_mb:.1f}MB) exceeds the maximum allowed size of 50MB.'
                })
                continue
            
            # If all checks pass, upload the file
            try:
                file_type = file_extension[1:] if file_extension else 'unknown'
                
                ProjectFile.objects.create(
                    project=project,
                    name=file_name,
                    file=file,
                    file_type=file_type,
                    size=file_size
                )
                successful_uploads += 1
                uploaded_files_names.append(file_name)
                
            except Exception as e:
                failed_uploads.append({
                    'name': file_name,
                    'reason': f'Upload failed: {str(e)}'
                })
        
        # Display appropriate messages
        if successful_uploads > 0:
            file_list = ', '.join(uploaded_files_names[:3])
            if len(uploaded_files_names) > 3:
                file_list += f' and {len(uploaded_files_names) - 3} more'
            
            create_project_version(
                project=project,
                user=request.user,
                action='file_added',
                description=f'Added {successful_uploads} file(s): {file_list}'
            )
            messages.success(request, f'✅ {successful_uploads} file(s) uploaded successfully!')

        if failed_uploads:
            for failed in failed_uploads:
                messages.error(request, f'❌ {failed["name"]}: {failed["reason"]}')
        
        if successful_uploads == 0 and not failed_uploads:
            messages.warning(request, 'No files were selected for upload.')
        
        return redirect('project_detail', project_id=project.id)
    
    return render(request, 'accounts/upload_file.html', {'project': project})

@login_required
def upload_version(request, project_id):
    """Upload a new version for a project"""
    from django.conf import settings
    
    project = get_object_or_404(Project, id=project_id)
    
    # Check if user has permission to upload versions
    if project.owner != request.user and request.user not in project.shared_with.all():
        messages.error(request, 'You do not have permission to upload versions to this project.')
        return redirect('project_detail', project_id=project.id)
    
    # Get next version number
    next_version = project.get_next_version_number()
    
    if request.method == 'POST':
        version_number = request.POST.get('version_number', '').strip()
        description = request.POST.get('description', '').strip()
        is_latest = request.POST.get('is_latest') == 'on'
        version_file = request.FILES.get('version_file')
        
        # Validation
        if not version_number:
            messages.error(request, 'Version number is required.')
            return redirect('upload_version', project_id=project.id)
        
        if not version_file:
            messages.error(request, 'Please select a file to upload.')
            return redirect('upload_version', project_id=project.id)
        
        # Check if version number already exists
        if ProjectVersion.objects.filter(project=project, version_number=version_number).exists():
            messages.error(request, f'Version {version_number} already exists. Please use a different version number.')
            return redirect('upload_version', project_id=project.id)
        
        # Validation constants
        MAX_FILE_SIZE = getattr(settings, 'MAX_FILE_SIZE', 50 * 1024 * 1024)  # 50MB
        BLOCKED_EXTENSIONS = getattr(settings, 'BLOCKED_FILE_EXTENSIONS', 
            ['.exe', '.bat', '.sh', '.cmd', '.com', '.app', '.dmg', '.deb', '.rpm'])
        
        file_size = version_file.size
        file_extension = os.path.splitext(version_file.name)[1].lower()
        
        # Security check - block dangerous file types
        if file_extension in BLOCKED_EXTENSIONS:
            messages.error(request, f'File type {file_extension} is not allowed for security reasons.')
            return redirect('upload_version', project_id=project.id)
        
        # File size check
        if file_size > MAX_FILE_SIZE:
            size_mb = file_size / (1024 * 1024)
            messages.error(request, f'File size ({size_mb:.1f}MB) exceeds the maximum allowed size of 50MB.')
            return redirect('upload_version', project_id=project.id)
        
        try:
            # If marking as latest, unmark all other versions
            if is_latest:
                ProjectVersion.objects.filter(project=project).update(is_latest=False)
            
            # Create the version
            version = ProjectVersion.objects.create(
                project=project,
                version_number=version_number,
                description=description if description else f'Version {version_number}',
                created_by=request.user,
                version_file=version_file,
                file_size=file_size,
                file_type=file_extension[1:] if file_extension else 'unknown',
                is_latest=is_latest,
                action='manual'
            )
            
            # Create files snapshot
            version.create_files_snapshot()
            
            # Update project timestamp
            project.save()
            
            messages.success(request, f'✅ Version {version_number} uploaded successfully!')
            return redirect('project_detail', project_id=project.id)
            
        except Exception as e:
            messages.error(request, f'Failed to upload version: {str(e)}')
            return redirect('upload_version', project_id=project.id)
    
    context = {
        'project': project,
        'next_version': next_version,
    }
    
    return render(request, 'accounts/upload_version.html', context)

@login_required
def my_projects(request):
    """List user's projects"""
    projects = Project.objects.filter(owner=request.user).order_by('-created_at')
    return render(request, 'accounts/my_projects.html', {'projects': projects})

# BROWSE PROJECTS 
@login_required
def browse_projects(request):
    """Browse all public projects with search and filter"""
    
    # Get search query from URL parameters
    search_query = request.GET.get('search', '')
    
    # Start with all public projects, excluding user's own projects
    projects = Project.objects.filter(is_public=True).exclude(owner=request.user).order_by('-updated_at')
    
    # Apply search filter if query exists
    if search_query:
        projects = projects.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(owner__first_name__icontains=search_query) |
            Q(owner__last_name__icontains=search_query) |
            Q(owner__username__icontains=search_query)
        )
    
    context = {
        'projects': projects,
        'search_query': search_query,
        'total_count': projects.count()
    }
    
    return render(request, 'accounts/browse_projects.html', context)

# SHARED WITH ME 
@login_required
def shared_with_me(request):
    """View projects shared with the current user"""
    
    # Get all projects where current user is in shared_with
    shared_projects = Project.objects.filter(
        shared_with=request.user
    ).order_by('-updated_at')
    
    context = {
        'projects': shared_projects,
        'total_count': shared_projects.count()
    }
    
    return render(request, 'accounts/shared_with_me.html', context)

# SHARE PROJECT 
@login_required
def share_project(request, project_id):
    """Share a project with other users"""
    project = get_object_or_404(Project, id=project_id, owner=request.user)
    
    if request.method == 'POST':
        # Get username or email to share with
        share_with = request.POST.get('share_with', '').strip()
        
        if not share_with:
            messages.error(request, 'Please enter a username or email.')
            return redirect('project_detail', project_id=project.id)
        
        try:
            # Find user by username or email
            user_to_share = User.objects.get(
                Q(username=share_with) | Q(email=share_with)
            )
            
            # Sharing not allowed with oneself
            if user_to_share == request.user:
                messages.error(request, 'You cannot share a project with yourself!')
                return redirect('project_detail', project_id=project.id)
            
            # Check if already shared
            if user_to_share in project.shared_with.all():
                messages.warning(request, f'Project is already shared with {user_to_share.username}!')
            else:
                project.shared_with.add(user_to_share)
                
                create_project_version(
                    project=project,
                    user=request.user,
                    action='shared',
                    description=f'Shared with {user_to_share.username}'
                )
                messages.success(request, f'Project shared with {user_to_share.username}!')
        
        except User.DoesNotExist:
            messages.error(request, f'User "{share_with}" not found!')
    
    return redirect('project_detail', project_id=project.id)


# UNSHARE PROJECT
@login_required
def unshare_project(request, project_id, user_id):
    """Remove a user from project sharing"""
    project = get_object_or_404(Project, id=project_id, owner=request.user)
    user_to_remove = get_object_or_404(User, id=user_id)
    
    # Remove user from shared_with
    project.shared_with.remove(user_to_remove)
    messages.success(request, f'Removed {user_to_remove.username} from project!')
    
    return redirect('project_detail', project_id=project.id)

# SECURITY QUESTION PASSWORD RESET
def password_reset_security(request):
    """Password reset using security question"""
    if request.method == 'POST':
        step = request.POST.get('step', '1')
        
        if step == '1':
            # Step 1: Verify username and show security question
            username = request.POST.get('username', '').strip()
            
            try:
                user = User.objects.get(username=username)
                
                if not user.security_question:
                    messages.error(request, 'This account does not have a security question set up.')
                    return redirect('password_reset')
                
                # Get the question text
                questions_dict = {
                    'pet': "What is your first pet's name?",
                    'city': "In what city were you born?",
                    'school': "What is your elementary school name?",
                    'color': "What is your favorite color?",
                    'food': "What is your favorite food?",
                }
                security_question = questions_dict.get(user.security_question, user.security_question)
                
                return render(request, 'accounts/password_reset_security.html', {
                    'step': 2,
                    'username': username,
                    'security_question': security_question
                })
                
            except User.DoesNotExist:
                messages.error(request, 'Username not found.')
                return render(request, 'accounts/password_reset_security.html', {'step': 1})
        
        elif step == '2':
            # Step 2: Verify answer and reset password
            username = request.POST.get('username')
            security_answer = request.POST.get('security_answer', '').lower().strip()
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            try:
                user = User.objects.get(username=username)
                
                if user.security_answer.lower().strip() == security_answer:
                    if new_password == confirm_password:
                        if len(new_password) >= 8:
                            user.set_password(new_password)
                            user.save()
                            messages.success(request, 'Password reset successful! You can now login with your new password.')
                            return redirect('login')
                        else:
                            messages.error(request, 'Password must be at least 8 characters long.')
                    else:
                        messages.error(request, 'Passwords do not match.')
                else:
                    messages.error(request, 'Incorrect answer to security question.')
                
                # Return to step 2 with error
                questions_dict = {
                    'pet': "What is your first pet's name?",
                    'city': "In what city were you born?",
                    'school': "What is your elementary school name?",
                    'color': "What is your favorite color?",
                    'food': "What is your favorite food?",
                }
                security_question = questions_dict.get(user.security_question, user.security_question)
                
                return render(request, 'accounts/password_reset_security.html', {
                    'step': 2,
                    'username': username,
                    'security_question': security_question
                })
                
            except User.DoesNotExist:
                messages.error(request, 'An error occurred. Please try again.')
                return redirect('password_reset')
    
    # GET request - show step 1
    return render(request, 'accounts/password_reset_security.html', {'step': 1})


# GENERATE SHAREABLE LINK
@login_required
def generate_share_link(request, project_id):
    """Generate or regenerate a shareable link for a project"""
    project = get_object_or_404(Project, id=project_id, owner=request.user)
    
    # If share_link doesn't exist, it will be auto-generated by the model
    # But we can also regenerate it if needed
    if request.method == 'POST' and request.POST.get('regenerate'):
        import uuid
        project.share_link = uuid.uuid4()
        project.save()
        messages.success(request, 'New shareable link generated!')
    
    return redirect('project_detail', project_id=project.id)


# ACCESS PROJECT VIA SHAREABLE LINK
def view_shared_project(request, share_uuid):
    """View a project via its shareable link"""
    try:
        project = get_object_or_404(Project, share_link=share_uuid)
        
        # Check if project is public or user has access
        if not project.is_public:
            if not request.user.is_authenticated:
                messages.error(request, 'This project is private. Please log in.')
                return redirect('login')
            
            if project.owner != request.user and request.user not in project.shared_with.all():
                messages.error(request, 'You do not have permission to view this private project.')
                return redirect('home')
        
        # Redirect to the project detail page
        return redirect('project_detail', project_id=project.id)
        
    except:
        messages.error(request, 'Invalid or expired share link.')
        return redirect('home')
    

# UPLOAD FOLDER
@login_required
def upload_folder(request):
    """Upload a folder from device and create a new project preserving folder structure"""
    if request.method == 'POST':
        files = request.FILES.getlist('files')
        is_public = request.POST.get('is_public') == 'on'
        
        if not files:
            messages.error(request, 'Please select a folder to upload.')
            return redirect('home')
        
        # Get folder name from first file's path
        first_file = files[0]
        folder_name = first_file.name.split('/')[0] if '/' in first_file.name else 'Uploaded Folder'
        
        # Create new project
        project = Project.objects.create(
            title=folder_name,
            description=f'Folder uploaded from device containing {len(files)} file(s)',
            owner=request.user,
            is_public=is_public
        )
        
        # Upload all files with folder structure preserved
        from django.conf import settings
        MAX_FILE_SIZE = getattr(settings, 'MAX_FILE_SIZE', 50 * 1024 * 1024)
        BLOCKED_EXTENSIONS = getattr(settings, 'BLOCKED_FILE_EXTENSIONS', 
            ['.exe', '.bat', '.sh', '.cmd', '.com', '.app', '.dmg', '.deb', '.rpm'])
        
        successful_uploads = 0
        failed_uploads = []
        
        for file in files:
            # Preserve the relative path
            file_name = file.name
            file_size = file.size
            file_extension = os.path.splitext(file_name)[1].lower()
            
            # Security checks
            if file_extension in BLOCKED_EXTENSIONS:
                failed_uploads.append(f'{file_name} (blocked file type)')
                continue
            
            if file_size > MAX_FILE_SIZE:
                failed_uploads.append(f'{file_name} (file too large)')
                continue
            
            try:
                file_type = file_extension[1:] if file_extension else 'unknown'
                
                # Store with relative path preserved
                ProjectFile.objects.create(
                    project=project,
                    name=file_name,  # This includes the folder path
                    file=file,
                    file_type=file_type,
                    size=file_size
                )
                successful_uploads += 1
            except Exception as e:
                failed_uploads.append(f'{file_name} (upload error: {str(e)})')
        
        # Create initial version
        create_project_version(
            project=project,
            user=request.user,
            action='created',
            description=f'Folder "{folder_name}" uploaded with {successful_uploads} file(s)'
        )
        
        # Show results
        if successful_uploads > 0:
            messages.success(request, f'✅ Folder "{folder_name}" uploaded with {successful_uploads} file(s)!')
        
        if failed_uploads:
            messages.warning(request, f'⚠️ {len(failed_uploads)} file(s) failed: {", ".join(failed_uploads[:3])}')
        
        return redirect('project_detail', project_id=project.id)
    
    return redirect('home')


# CREATE DOCUMENT
@login_required
def create_document(request):
    """Create a new document (Google Docs style)"""
    if request.method == 'POST':
        doc_title = request.POST.get('title', '').strip()
        doc_content = request.POST.get('content', '')
        is_public = request.POST.get('is_public') == 'on'
        
        if not doc_title:
            doc_title = f'Untitled Document {timezone.now().strftime("%Y%m%d_%H%M%S")}'
        
        # Create project for the document
        project = Project.objects.create(
            title=doc_title,
            description='Text document',
            owner=request.user,
            is_public=is_public
        )
        
        # Create a text file with the content
        from django.core.files.base import ContentFile
        
        file_content = doc_content if doc_content else '# ' + doc_title + '\n\nStart writing here...'
        file_name = f'{doc_title}.md'
        
        ProjectFile.objects.create(
            project=project,
            name=file_name,
            file=ContentFile(file_content.encode('utf-8'), name=file_name),
            file_type='md',
            size=len(file_content.encode('utf-8'))
        )
        
        messages.success(request, f'✅ Document "{doc_title}" created successfully!')
        return redirect('project_detail', project_id=project.id)
    
    return render(request, 'accounts/create_document.html')

# View version details
@login_required
def view_version(request, project_id, version_id):
    """View details of a specific version"""
    project = get_object_or_404(Project, id=project_id)
    version = get_object_or_404(ProjectVersion, id=version_id, project=project)
    
    # Check permissions
    if not project.is_public and project.owner != request.user and request.user not in project.shared_with.all():
        messages.error(request, 'You do not have permission to view this project.')
        return redirect('home')
    
    # Get files from snapshot
    snapshot_files = version.files_snapshot.get('files', [])
    
    return render(request, 'accounts/view_version.html', {
        'project': project,
        'version': version,
        'snapshot_files': snapshot_files,
    })


# Version history page
@login_required
def version_history(request, project_id):
    """View complete version history for a project"""
    project = get_object_or_404(Project, id=project_id)
    
    # Check permissions
    if not project.is_public and project.owner != request.user and request.user not in project.shared_with.all():
        messages.error(request, 'You do not have permission to view this project.')
        return redirect('home')
    
    versions = project.versions.all()
    
    return render(request, 'accounts/version_history.html', {
        'project': project,
        'versions': versions,
    })


# Restore version
@login_required
def restore_version(request, project_id, version_id):
    """Restore project to a specific version"""
    project = get_object_or_404(Project, id=project_id)
    version = get_object_or_404(ProjectVersion, id=version_id, project=project)
    
    # Only owner can restore
    if project.owner != request.user:
        messages.error(request, 'Only the project owner can restore versions.')
        return redirect('project_detail', project_id=project.id)
    
    if request.method == 'POST':
        try:
            create_project_version(
                project=project,
                user=request.user,
                action='restored', 
                description=f'🔄 Restored to {version.version_number} from {version.created_at.strftime("%Y-%m-%d %H:%M")}. Original description: {version.description}'
            )
            
            messages.success(request, f'✅ Project restored to {version.version_number}!')
            
        except Exception as e:
            messages.error(request, f'Failed to restore version: {str(e)}')
    
    return redirect('project_detail', project_id=project.id)

# DELETE FILE
@login_required
def delete_file(request, project_id, file_id):
    """Delete a file from project"""
    project = get_object_or_404(Project, id=project_id)
    file = get_object_or_404(ProjectFile, id=file_id, project=project)
    
    # Check permissions
    can_delete = False
    
    # Owner can always delete
    if project.owner == request.user:
        can_delete = True
    # Collaborators can delete if owner allows it
    elif request.user in project.shared_with.all() and project.allow_collaborators_delete:
        can_delete = True
    
    if not can_delete:
        messages.error(request, 'You do not have permission to delete files from this project.')
        return redirect('project_detail', project_id=project.id)
    
    if request.method == 'POST':
        file_name = file.name
        
        # Delete the physical file
        if file.file:
            file.file.delete()
        
        # Delete the database record
        file.delete()
        
        # Create version for file deletion
        create_project_version(
            project=project,
            user=request.user,
            action='file_deleted',
            description=f'Deleted file: {file_name}'
        )
        
        messages.success(request, f'File "{file_name}" deleted successfully!')
        return redirect('project_detail', project_id=project.id)
    
    # if GET request, show confirmation modal 
    return redirect('project_detail', project_id=project.id)


# TOGGLE DELETE PERMISSION
@login_required
def toggle_delete_permission(request, project_id):
    """Toggle collaborator delete permission"""
    project = get_object_or_404(Project, id=project_id, owner=request.user)
    
    if request.method == 'POST':
        project.allow_collaborators_delete = not project.allow_collaborators_delete
        project.save()
        
        status = "enabled" if project.allow_collaborators_delete else "disabled"
        messages.success(request, f'Collaborator file deletion {status}!')
    
    return redirect('project_detail', project_id=project.id)

# VIEW/EDIT FILE
@login_required
def view_edit_file(request, project_id, file_id):
    """View and edit file content"""
    project = get_object_or_404(Project, id=project_id)
    file = get_object_or_404(ProjectFile, id=file_id, project=project)
    
    # Check permissions
    if not project.is_public and project.owner != request.user and request.user not in project.shared_with.all():
        messages.error(request, 'You do not have permission to view this file.')
        return redirect('project_detail', project_id=project.id)
    
    # Define editable file types
    TEXT_EXTENSIONS = [
        'txt', 'md', 'markdown', 'html', 'htm', 'css', 'js', 'json', 
        'py', 'java', 'cpp', 'c', 'h', 'xml', 'yml', 'yaml', 'ini', 
        'cfg', 'conf', 'log', 'csv', 'sql', 'sh', 'bat', 'php', 'rb',
        'go', 'rs', 'ts', 'jsx', 'tsx', 'vue', 'swift', 'kt', 'r'
    ]
    WORD_EXTENSIONS = ['docx', 'doc']
    EXCEL_EXTENSIONS = ['xlsx', 'xls']
    
    file_extension = file.file_type.lower()
    file_type = 'text'  # Default
    file_content = None
    excel_data = None
    is_editable = False
    
    # Import BytesIO for memory operations
    from io import BytesIO
    
    # Determine file type and read content
    if file_extension in TEXT_EXTENSIONS:
        file_type = 'text'
        is_editable = True
        file_content = ''
        
        try:
            # Read file into memory
            file.file.seek(0)
            file_bytes = file.file.read()
            file_content = file_bytes.decode('utf-8', errors='ignore')
            print(f"✅ Text file loaded: {len(file_content)} chars")
            
        except Exception as e:
            print(f"⚠️ Text file error: {str(e)}")
            import traceback
            print(traceback.format_exc())
            messages.warning(request, 'Could not read file content. Starting with empty editor.')
            file_content = ''
    
    elif file_extension in WORD_EXTENSIONS:
        file_type = 'word'
        is_editable = True
        file_content = ''
        
        try:
            import mammoth
            
            # Read file into memory
            file.file.seek(0)
            file_bytes = file.file.read()
            file_stream = BytesIO(file_bytes)
            
            result = mammoth.convert_to_html(file_stream)
            file_content = result.value
            print(f"✅ Word file loaded successfully")
            
        except ImportError:
            print("⚠️ mammoth library not installed")
            messages.warning(request, 'Word file support not available. Starting with empty editor.')
            file_content = '<p>Document content could not be loaded.</p>'
            
        except Exception as e:
            print(f"⚠️ Word file error: {str(e)}")
            import traceback
            print(traceback.format_exc())
            messages.warning(request, 'Could not read Word file. Starting with empty editor.')
            file_content = '<p>Document content could not be loaded.</p>'
    
    elif file_extension in EXCEL_EXTENSIONS:
        file_type = 'excel'
        is_editable = True
        excel_data = {'headers': [], 'rows': []}
        
        try:
            import openpyxl
            
            # Read file into memory
            file.file.seek(0)
            file_bytes = file.file.read()
            file_stream = BytesIO(file_bytes)
            
            wb = openpyxl.load_workbook(file_stream, data_only=True)
            sheet = wb.active
            
            # Convert to list of lists
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    excel_data['headers'] = [str(cell) if cell is not None else '' for cell in row]
                else:
                    excel_data['rows'].append([str(cell) if cell is not None else '' for cell in row])
            
            wb.close()
            print(f"✅ Excel file loaded: {len(excel_data['rows'])} rows")
            
        except ImportError:
            print("⚠️ openpyxl library not installed")
            messages.warning(request, 'Excel file support not available. Starting with empty spreadsheet.')
            excel_data = {
                'headers': ['Column 1', 'Column 2', 'Column 3'],
                'rows': [['', '', ''], ['', '', '']]
            }
            
        except Exception as e:
            print(f"⚠️ Excel file error: {str(e)}")
            import traceback
            print(traceback.format_exc())
            messages.warning(request, 'Could not read Excel file. Starting with empty spreadsheet.')
            excel_data = {
                'headers': ['Column 1', 'Column 2', 'Column 3'],
                'rows': [['', '', ''], ['', '', '']]
            }
    
    # Handle file save (POST request)
    if request.method == 'POST' and is_editable:
        # Check if user can edit
        can_edit = project.owner == request.user or request.user in project.shared_with.all()
        
        if not can_edit:
            messages.error(request, 'You do not have permission to edit this file.')
            return redirect('view_edit_file', project_id=project.id, file_id=file.id)
        
        try:
            if file_type == 'text':
                # Save text file
                new_content = request.POST.get('content', '')
                from django.core.files.base import ContentFile
                file.file.save(file.name, ContentFile(new_content.encode('utf-8')), save=False)
                file.size = len(new_content.encode('utf-8'))
                file.save()
                file_content = new_content
            
            elif file_type == 'word':
                # Save Word file
                from docx import Document
                
                new_content = request.POST.get('content', '')
                
                # Create new document
                doc = Document()
                
                # Parse HTML content and add to document (simple conversion)
                import re
                clean_text = re.sub('<[^<]+?>', '', new_content)
                
                for paragraph in clean_text.split('\n'):
                    if paragraph.strip():
                        doc.add_paragraph(paragraph)
                
                # Save to BytesIO
                doc_buffer = BytesIO()
                doc.save(doc_buffer)
                doc_buffer.seek(0)
                
                from django.core.files.base import ContentFile
                file.file.save(file.name, ContentFile(doc_buffer.read()), save=False)
                file.size = doc_buffer.tell()
                file.save()
                
                # Re-read for display
                import mammoth
                file.file.seek(0)
                file_bytes = file.file.read()
                file_stream = BytesIO(file_bytes)
                result = mammoth.convert_to_html(file_stream)
                file_content = result.value
            
            elif file_type == 'excel':
                # Save Excel file
                import openpyxl
                import json
                
                # Get JSON data from form
                excel_json = request.POST.get('excel_data', '')
                data = json.loads(excel_json)
                
                # Create new workbook
                wb = openpyxl.Workbook()
                sheet = wb.active
                
                # Write headers
                if data.get('headers'):
                    sheet.append(data['headers'])
                
                # Write rows
                for row in data.get('rows', []):
                    sheet.append(row)
                
                # Save to BytesIO
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                from django.core.files.base import ContentFile
                file.file.save(file.name, ContentFile(excel_buffer.read()), save=False)
                file.size = excel_buffer.tell()
                file.save()
                
                # Re-read for display
                file.file.seek(0)
                file_bytes = file.file.read()
                file_stream = BytesIO(file_bytes)
                wb = openpyxl.load_workbook(file_stream, data_only=True)
                sheet = wb.active
                excel_data = {'headers': [], 'rows': []}
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i == 0:
                        excel_data['headers'] = [str(cell) if cell is not None else '' for cell in row]
                    else:
                        excel_data['rows'].append([str(cell) if cell is not None else '' for cell in row])
                wb.close()
            
            # Create version for file edit
            create_project_version(
                project=project,
                user=request.user,
                action='file_updated',
                description=f'Updated file: {file.name}'
            )
            
            messages.success(request, f'File "{file.name}" saved successfully!')
            
        except Exception as e:
            print(f"❌ Save error: {str(e)}")
            import traceback
            print(traceback.format_exc())
            messages.error(request, f'Failed to save file: {str(e)}')
    
    context = {
        'project': project,
        'file': file,
        'file_type': file_type,
        'is_editable': is_editable,
        'file_content': file_content,
        'excel_data': excel_data,
        'can_edit': project.owner == request.user or request.user in project.shared_with.all(),
    }
    
    return render(request, 'accounts/view_edit_file.html', context)

# MOVE TO TRASH
@login_required
def move_to_trash(request, project_id):
    """Move project to trash (soft delete)"""
    project = get_object_or_404(Project, id=project_id, owner=request.user)
    
    if request.method == 'POST':
        project.is_deleted = True
        project.deleted_at = timezone.now()
        project.save()
        
        messages.success(request, f'Project "{project.title}" moved to trash!')
        return redirect('home')
    
    return redirect('home')


# VIEW TRASH
@login_required
def trash(request):
    """View deleted projects"""
    deleted_projects = Project.objects.filter(
        owner=request.user,
        is_deleted=True
    ).order_by('-deleted_at')
    
    context = {
        'projects': deleted_projects,
    }
    
    return render(request, 'accounts/trash.html', context)


# RESTORE FROM TRASH
@login_required
def restore_from_trash(request, project_id):
    """Restore project from trash"""
    project = get_object_or_404(Project, id=project_id, owner=request.user, is_deleted=True)
    
    if request.method == 'POST':
        project.is_deleted = False
        project.deleted_at = None
        project.save()
        
        messages.success(request, f'Project "{project.title}" restored!')
        return redirect('trash')
    
    return redirect('trash')


# PERMANENTLY DELETE
@login_required
def delete_permanently(request, project_id):
    """Permanently delete project from trash"""
    project = get_object_or_404(Project, id=project_id, owner=request.user, is_deleted=True)
    
    if request.method == 'POST':
        project_title = project.title
        
        # Delete all associated files
        for file in project.files.all():
            if file.file:
                file.file.delete()
            file.delete()
        
        # Delete the project
        project.delete()
        
        messages.success(request, f'Project "{project_title}" permanently deleted!')
        return redirect('trash')
    
    return redirect('trash')


# EMPTY TRASH
@login_required
def empty_trash(request):
    """Empty entire trash"""
    if request.method == 'POST':
        deleted_projects = Project.objects.filter(
            owner=request.user,
            is_deleted=True
        )
        
        count = deleted_projects.count()
        
        # Delete all files and projects
        for project in deleted_projects:
            for file in project.files.all():
                if file.file:
                    file.file.delete()
                file.delete()
            project.delete()
        
        messages.success(request, f'Trash emptied! {count} project(s) permanently deleted.')
        return redirect('trash')
    
    return redirect('trash')